#!/usr/bin/python
# -*- coding: utf-8 -*-

"""
__title__=
__Time__= 2021/1/26 10:02
__Author__=
__Des__=
"""

import openpyxl
from openpyxl.styles import Font
from comn.init_path import BASHPATH
import json
import time
import os

class Oper_Excel(object):
    def __init__(self,filename=None):
        if filename:
            # self.xls_wb= openpyxl.load_workbook(filename)
            self.filename=os.path.join(BASHPATH, "test_data", filename)
            self.xls_wb = openpyxl.load_workbook(self.filename)
        else:
            self.xls_wb= openpyxl.Workbook()

        self.sheet=self.get_sheet()


    #对行进行遍历，输出单元格的值
    def get_sheet(self):
        '''
        获取打开时的当前表单
        :return:
        '''
        return self.xls_wb.active
        # return self.xls_wb.get_sheet_by_name('api')

    def get_sheet_title(self):
        '''
        获取表单名称
        :return:
        '''
        return self.sheet.title

    def get_cases(self):
        '''
        读取xlsx表格内容
        :return:
        '''
        row_nums=self.sheet.max_row
        header_title=[title for title in self.get_rows_value(1)]
        casedatas=[]
        #获取表头以外数据
        for row in range(2,row_nums+1):
            data=[cell_value for cell_value in self.get_rows_value(row)]
            case=dict(zip(header_title,data))
            if case["isrun"]:
                casedatas.append(case)
        return casedatas

    def get_columns_value(self,key):
        '''
        获取某一列的数据
        :param key:
        :return:
        '''
        col_values=[]
        for cell in self.sheet[key]:
            col_values.append(cell.value)
        return col_values

    def get_rows_value(self,row):
        '''
        获取某一行的值
        :param row:
        :return:
        '''
        row_values=[]
        for cell in self.sheet[row]:
            row_values.append(cell.value)
        return row_values

    def set_cellstyle(self,row,col):
        '''
        设置单元格字体样式
        :param row:
        :param col:
        :return:
        '''
        font=Font(name='微软雅黑',
                  size=11,
                  color='FF000000')
        self.sheet.cell(row,col).font=font


    def get_cell_value(self,row,col):
        '''
        获取单元格的值
        :param row:
        :param col:
        :return:
        '''
        return self.sheet.cell(row,col).value

    def set_cell_value(self,row,col,value):
        '''
        设置单元格的值
        :param row:
        :param col:
        :param value:
        :return:
        '''
        self.sheet.cell(row,col,value)
        self.xls_wb.save(self.filename)

    def get_row_index(self,key,value):
        '''
        获取行索引
        :param key:
        :param value:
        :return:
        '''
        cellsvalue=self.get_columns_value(key)
        for num,cvalue in enumerate(cellsvalue):
            if cvalue==value:
                return num+1

    def write_xlsx(self,list):
        '''
        将列表内容写入xlsx
        :param list:
        :return:
        '''

        row_nums=self.sheet.max_row+1
        self.sheet.delete_rows(2,row_nums)
        for li in list:
            self.sheet.append(li)
        for row in range(1, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column + 1):
                self.set_cellstyle(row, col)
        # self.xls_wb.save(time.strftime("%Y%m%d%H%M%S")+'.xlsx')
        self.xls_wb.save(self.filename)



if __name__ == '__main__':
    xls=Oper_Excel("fsd_api.xlsx")
    # print(xls.get_row_index("A","FSD0001"))
    print(xls.get_cell_value(xls.get_row_index("A", "FSD0001"), 15))